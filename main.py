# https://cloud.google.com/docs/authentication/getting-started#auth-cloud-implicit-python

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Side, Color

from service_cloud import analyze_entities

FILL1 = PatternFill(fill_type='solid', fgColor= Color(rgb='90EE90'))
FILL2 = PatternFill(fill_type='solid', fgColor= Color(rgb='006400'))
FILL3 = PatternFill(fill_type='solid', fgColor= Color(rgb='FFFF00'))
FILL4 = PatternFill(fill_type='solid', fgColor= Color(rgb='CD853F'))
FILL5 = PatternFill(fill_type='solid', fgColor= Color(rgb='cfb53b'))
FILL_TITLE = PatternFill(fill_type='solid', fgColor= Color(rgb='42aaff'))

COLORS = {
    'OTHER':'90EE90',
    'PERSON':'42aaff',
    'CONSUMER_GOOD':'FFFF00',
    'EVENT':'CD853F',
    'NUMBER':'cfb53b',
}
def process_entities():
    wb = load_workbook('table.xlsx')
    ws = wb['text']
    base = ws.values
    next(base)
    result = list()
    for row in base:
        for result_analyze in analyze_entities(row[1]):
            r = list(result_analyze.values())
            r.insert(0, row[0])
            r.insert(1, row[1])
            result.append(r)
    ws = wb.create_sheet('result')
    ws.append(['Name', 'Text', 'entity', 'score', 'type'])
    #ws['A1'].fill = FILL1
    #ws['B1'].fill = FILL2
    #ws['C1'].fill = FILL3
    #ws['D1'].fill = FILL4
    #ws['E1'].fill = FILL5
    for row in result:
        ws.append(row)
    for i in range(2, len(result) + 2):
        ws[f'E{i}'].fill = PatternFill(fill_type='solid', fgColor= Color(rgb=COLORS.get(ws[f'E{i}'].value, '42aaff')))
    wb.save('table.xlsx')

def save(base):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Результат'
    ws.append(['Name', 'Text', 'entity', 'score', 'type'])
    for row in base:
        ws.append(row)
    wb.save('result.xlsx')

def main():
    with open('text.txt', 'r', encoding='utf-8') as f: text_content = f.read()
    #analyze_entities(text_content)
    process_entities()


if __name__ == '__main__':
    main()
    